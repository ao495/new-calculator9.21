# -*- coding: utf-8 -*-

import os
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, Menu
from PIL import Image, ImageTk, ImageDraw
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import io
import json

class PhotoLayoutApp:
    """
    Excelに写真をレイアウトするGUIアプリケーション
    全機能統合版 (修正済み)
    """
    def __init__(self, root):
        self.root = root
        self.root.title("Excel写真レイアウトアプリ")
        
        # --- 状態変数 ---
        self.photos = []  # PIL Imageオブジェクト
        self.photo_paths = []  # 写真のパス
        self.rows_config = [0, 0]  # 各段の列数
        self.main_indices = [0, 0]  # 各段のメイン写真のインデックス
        self.row_comments = ["", ""]  # 各段のコメント
        self.photo_comments = {}  # 個別写真のコメント
        
        # --- UI表示用の変数とオブジェクト ---
        self.photo_positions = []  # Canvas上の写真の位置情報
        self.sliders = []  # メイン比率スライダー
        self.height_sliders = []  # 高さ倍率スライダー
        self.comment_entries = []  # 段ごとのコメント入力欄
        self.tk_images = []  # Canvasで表示するためのTkImageオブジェクト
        
        # --- ドラッグ＆ドロップ用の状態 ---
        self.drag_data = {"item": None, "photo_idx": None, "x": 0, "y": 0}
        
        # --- 範囲選択モードの状態 ---
        self.select_mode = False
        self.select_start = None
        
        self.setup_ui()
        self.update_sliders()
        
        # ウィンドウを全画面表示
        self.root.state('zoomed')
        self.root.update()
        self.update_preview()

    # --- UI構築 ---
    def setup_ui(self):
        """
        GUIの各ウィジェットを構築する
        """
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill="both", expand=True)
        
        # ボタンフレーム
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=5, padx=5)
        
        tk.Button(btn_frame, text="画像追加", command=self.add_photos, width=12).pack(side="left", padx=2)
        tk.Button(btn_frame, text="写真削除", command=self.remove_selected_photo, width=12).pack(side="left", padx=2)
        tk.Button(btn_frame, text="写真入れ替え", command=self.swap_photos_dialog, width=12).pack(side="left", padx=2)
        tk.Button(btn_frame, text="コメント追加", command=self.toggle_select_mode, width=12).pack(side="left", padx=2)
        tk.Button(btn_frame, text="Excel出力", command=self.export_excel, width=12).pack(side="left", padx=2)
        tk.Button(btn_frame, text="設定保存", command=self.save_config, width=12).pack(side="left", padx=2)
        tk.Button(btn_frame, text="設定読み込み", command=self.load_config, width=12).pack(side="left", padx=2)
        tk.Button(btn_frame, text="ヘルプ", command=self.show_help, width=12).pack(side="left", padx=2)
        
        # キャンバスとスクロールバー
        canvas_frame = tk.Frame(main_frame)
        canvas_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.canvas = tk.Canvas(canvas_frame, bg="#eee")
        self.canvas.pack(side="left", fill="both", expand=True)
        
        scrollbar_y = tk.Scrollbar(canvas_frame, orient="vertical", command=self.canvas.yview)
        scrollbar_y.pack(side="right", fill="y")
        self.canvas.configure(yscrollcommand=scrollbar_y.set)
        
        scrollbar_x = tk.Scrollbar(main_frame, orient="horizontal", command=self.canvas.xview)
        scrollbar_x.pack(side="bottom", fill="x")
        self.canvas.configure(xscrollcommand=scrollbar_x.set)

        self.canvas.bind("<ButtonPress-1>", self.on_press)
        self.canvas.bind("<B1-Motion>", self.on_motion)
        self.canvas.bind("<ButtonRelease-1>", self.on_release)
        # 右クリックイベントをバインド
        self.canvas.bind("<Button-3>", self.on_right_click)
        self.canvas.bind("<Configure>", lambda e: self.update_preview())
        self.root.bind("<Delete>", self.on_delete_key)
        self.root.bind("<BackSpace>", self.on_delete_key)

        self.slider_frame = tk.Frame(main_frame)
        self.slider_frame.pack(fill="x", padx=5, pady=5)
        
        self.status_var = tk.StringVar()
        status_bar = tk.Label(main_frame, textvariable=self.status_var, anchor="w", bd=1, relief=tk.SUNKEN)
        status_bar.pack(fill="x", side="bottom")

    # --- スライダー更新 ---
    def update_sliders(self):
        """
        段数設定に合わせてスライダーUIを更新する
        """
        for w in self.slider_frame.winfo_children():
            w.destroy()
        self.sliders.clear()
        self.height_sliders.clear()
        self.comment_entries.clear()
        
        num_rows = 2
        for i in range(num_rows):
            frame = tk.LabelFrame(self.slider_frame, text=f"段 {i+1} 設定")
            frame.pack(side="left", padx=5, pady=5, expand=True, fill="x")
            
            # メイン比率スライダー
            tk.Label(frame, text="メイン比率").pack()
            slider = tk.Scale(frame, from_=0.1, to=0.9, resolution=0.01, orient="horizontal",
                             command=lambda e, idx=i: self.update_preview())
            slider.set(0.5)
            slider.pack(fill="x")
            self.sliders.append(slider)

            # 高さ倍率スライダー
            tk.Label(frame, text="高さ倍率").pack()
            h_slider = tk.Scale(frame, from_=0.1, to=2.0, resolution=0.01, orient="horizontal",
                                command=lambda e, idx=i: self.update_preview())
            h_slider.set(0.5)
            h_slider.pack(fill="x")
            self.height_sliders.append(h_slider)
            
            # 段ごとのコメント入力欄
            tk.Label(frame, text="段コメント").pack()
            comment_var = tk.StringVar(value=self.row_comments[i] if i < len(self.row_comments) else "")
            comment_entry = tk.Entry(frame, textvariable=comment_var)
            comment_entry.pack(fill="x")
            self.comment_entries.append(comment_entry)
            
            comment_entry.bind("<KeyRelease>", lambda e, idx=i: self.update_row_comment(idx, comment_entry.get()))

    def update_row_comment(self, row_idx, text):
        """
        段コメント入力欄の値が変更されたときに呼び出される
        """
        while len(self.row_comments) <= row_idx:
            self.row_comments.append("")
        self.row_comments[row_idx] = text
        self.update_preview()

    def toggle_select_mode(self):
        """
        コメント追加モードのON/OFFを切り替える
        """
        self.select_mode = not self.select_mode
        if self.select_mode:
            self.status_var.set("コメント追加モード: 写真を範囲選択してください")
            self.canvas.config(cursor="cross")
            self.drag_data = {"item": None, "photo_idx": None, "x": 0, "y": 0}
        else:
            self.status_var.set("通常モード")
            self.canvas.config(cursor="arrow")
            self.canvas.delete("selection_box")

    # --- 写真追加・削除 ---
    def add_photos(self):
        """
        ファイルダイアログから画像ファイルを選択し、リストに追加し、自動で振り分ける
        """
        files = filedialog.askopenfilenames(filetypes=[
            ("画像ファイル", "*.png;*.jpg;*.jpeg;*.bmp;*.gif;*.tiff"),
            ("すべてのファイル", "*.*")
        ])
        
        if not files:
            return
            
        added = 0
        for f in files:
            try:
                img = Image.open(f)
                self.photos.append(img)
                self.photo_paths.append(f)
                added += 1
            except Exception as e:
                self.status_var.set(f"画像読み込み失敗: {os.path.basename(f)} ({str(e)})")
        
        if added > 0:
            self.auto_distribute_photos()
            self.update_preview()

    def auto_distribute_photos(self):
        """
        写真の枚数に基づいて、1段目と2段目に自動で振り分ける
        条件: 4枚目までを上段、それ以降を下段
        """
        total_photos = len(self.photos)
        
        if total_photos <= 4:
            self.rows_config = [total_photos, 0]
        else:
            self.rows_config = [4, total_photos - 4]
            
        for i in range(len(self.main_indices)):
            if self.rows_config[i] > 0:
                self.main_indices[i] = min(self.main_indices[i], self.rows_config[i] - 1)
            else:
                self.main_indices[i] = 0

    def remove_selected_photo(self):
        """
        選択中の写真をリストから削除し、再度自動振り分けを行う
        """
        if self.drag_data["photo_idx"] is not None and 0 <= self.drag_data["photo_idx"] < len(self.photos):
            idx = self.drag_data["photo_idx"]
            del self.photos[idx]
            del self.photo_paths[idx]
            
            new_photo_comments = {}
            for key, value in self.photo_comments.items():
                if key < idx:
                    new_photo_comments[key] = value
                elif key > idx:
                    new_photo_comments[key - 1] = value
            self.photo_comments = new_photo_comments
            
            self.drag_data = {"item": None, "photo_idx": None, "x": 0, "y": 0}
            self.auto_distribute_photos()
            self.update_preview()
            self.status_var.set("選択した写真を削除しました")
        else:
            messagebox.showinfo("情報", "削除する写真を選択してください")

    def swap_photos_dialog(self):
        """
        番号指定で写真の位置を入れ替えるダイアログを表示する
        """
        if len(self.photos) < 2:
            messagebox.showinfo("情報", "写真を2枚以上追加してください。")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("写真入れ替え")
        dialog.geometry("300x150")

        tk.Label(dialog, text="入れ替えたい写真の番号を入力してください。").pack(pady=5)
        tk.Label(dialog, text="(例: 1, 3)").pack()

        frame = tk.Frame(dialog)
        frame.pack(pady=5)
        
        tk.Label(frame, text="番号1:").pack(side="left", padx=5)
        entry1 = tk.Entry(frame, width=5)
        entry1.pack(side="left", padx=5)
        
        tk.Label(frame, text="番号2:").pack(side="left", padx=5)
        entry2 = tk.Entry(frame, width=5)
        entry2.pack(side="left", padx=5)

        def perform_swap():
            try:
                idx1 = int(entry1.get()) - 1
                idx2 = int(entry2.get()) - 1

                if not (0 <= idx1 < len(self.photos) and 0 <= idx2 < len(self.photos)):
                    messagebox.showerror("エラー", "無効な写真番号です。")
                    return

                self.photos[idx1], self.photos[idx2] = self.photos[idx2], self.photos[idx1]
                self.photo_paths[idx1], self.photo_paths[idx2] = self.photo_paths[idx2], self.photo_paths[idx1]
                
                comment_src = self.photo_comments.get(idx1)
                comment_dst = self.photo_comments.get(idx2)
                
                self.photo_comments.pop(idx1, None)
                self.photo_comments.pop(idx2, None)
                if comment_src:
                    self.photo_comments[idx2] = comment_src
                if comment_dst:
                    self.photo_comments[idx1] = comment_dst
                
                # 自動振り分けを呼び出さないように修正
                self.update_preview()
                self.status_var.set(f"写真 {idx1+1} と {idx2+1} を入れ替えました")
                dialog.destroy()
            except ValueError:
                messagebox.showerror("エラー", "有効な数字を入力してください。")

        tk.Button(dialog, text="入れ替え実行", command=perform_swap).pack(pady=10)

    # --- Canvas描画 ---
    def update_preview(self):
        """
        設定に基づいてCanvas上に写真のプレビューを描画する
        画面幅に合わせて画像サイズを調整
        """
        self.canvas.delete("all")
        self.photo_positions.clear()
        self.tk_images = []
        
        if not self.photos:
            self.status_var.set("写真を追加してください")
            return
            
        canvas_w = self.canvas.winfo_width()
        if canvas_w < 100:
            canvas_w = self.root.winfo_width() - 50
        
        y = 10
        idx = 0
        total_width = 0
        
        for r_idx, count in enumerate(self.rows_config):
            if idx >= len(self.photos):
                break
            
            row_photos = self.photos[idx:idx + count]
            if not row_photos:
                break
            
            row_comment_text = self.row_comments[r_idx] if r_idx < len(self.row_comments) else ""
            if row_comment_text:
                self.canvas.create_text(canvas_w // 2, y, text=row_comment_text, anchor="n", fill="black", font=("Arial", 12, "bold"))
                y += 25
            
            main_idx = self.main_indices[r_idx] if r_idx < len(self.main_indices) else 0
            if main_idx >= len(row_photos):
                main_idx = 0
                
            main_ratio = self.sliders[r_idx].get()
            photo_count = len(row_photos)
            
            if photo_count == 1:
                main_ratio = 1.0
            remaining_ratio = (1 - main_ratio) / (photo_count - 1) if photo_count > 1 else 0
            
            widths = []
            for i in range(photo_count):
                w = int(canvas_w * (main_ratio if i == main_idx else remaining_ratio))
                widths.append(w)

            max_height = 0
            for i, img in enumerate(row_photos):
                h = int(img.height * (widths[i] / img.width))
                max_height = max(max_height, h)
            
            height_ratio = self.height_sliders[r_idx].get() if r_idx < len(self.height_sliders) else 0.5
            uniform_height = max_height * height_ratio

            x = 0
            row_positions = []
            for i, img in enumerate(row_photos):
                w = widths[i]
                h = int(img.height * (w / img.width))
                
                scale_factor = uniform_height / h
                new_w = int(w * scale_factor)
                
                if i == photo_count - 1:
                    new_w = canvas_w - x
                
                try:
                    resized_img = img.resize((new_w, int(uniform_height)), Image.LANCZOS)
                    tk_img = ImageTk.PhotoImage(resized_img)
                    self.tk_images.append(tk_img)
                    
                    self.canvas.create_image(x, y, anchor="nw", image=tk_img,
                                             tags=(f"photo_{r_idx}_{i}", "photo_item"))
                    
                    if i == main_idx:
                        self.canvas.create_rectangle(x, y, x + new_w, y + uniform_height, outline="red", width=3,
                                                     tags=(f"main_{r_idx}",))
                    
                    photo_idx = sum(self.rows_config[:r_idx]) + i
                    if photo_idx in self.photo_comments:
                        self.canvas.create_text(x + new_w//2, y + 10, text=self.photo_comments[photo_idx],
                                                anchor="n", fill="blue", font=("Arial", 10))

                    if photo_idx < len(self.photo_paths):
                        filename = os.path.basename(self.photo_paths[photo_idx])
                        text_to_show = f"({photo_idx+1}) {filename[:20]}"
                        self.canvas.create_text(x + new_w//2, y + uniform_height + 5, text=text_to_show,
                                                 anchor="n", fill="black", tags=(f"label_{r_idx}_{i}",))
                    
                    row_positions.append((x, y, x + new_w, y + uniform_height))
                    x += new_w
                except Exception as e:
                    self.status_var.set(f"画像表示エラー: {str(e)}")
            
            self.photo_positions.append(row_positions)
            y += uniform_height + 30
            idx += count
            total_width = max(total_width, x)
        
        self.canvas.config(scrollregion=(0, 0, total_width, y))
        self.status_var.set(f"表示中: {len(self.photos)}枚の写真")

    # --- ドラッグ＆ドロップ操作 ---
    def on_press(self, event):
        """
        マウスボタンが押されたときの処理
        """
        self.canvas.delete("selection")
        if self.select_mode:
            self.select_start = (self.canvas.canvasx(event.x), self.canvas.canvasy(event.y))
            self.canvas.delete("selection_box")
        else:
            photo_idx_found = None
            for r, row in enumerate(self.photo_positions):
                for i, (x1, y1, x2, y2) in enumerate(row):
                    if x1 <= self.canvas.canvasx(event.x) <= x2 and y1 <= self.canvas.canvasy(event.y) <= y2:
                        photo_idx = sum(self.rows_config[:r]) + i
                        if photo_idx < len(self.photos):
                            photo_idx_found = photo_idx
                            break
                if photo_idx_found is not None:
                    break
            
            if photo_idx_found is not None:
                self.drag_data = {
                    "item": self.canvas.find_withtag(f"photo_{r}_{i}")[0],
                    "photo_idx": photo_idx_found,
                    "x": self.canvas.canvasx(event.x),
                    "y": self.canvas.canvasy(event.y),
                    "row": r,
                    "idx": i
                }
                x1, y1, x2, y2 = self.photo_positions[self.drag_data["row"]][self.drag_data["idx"]]
                self.canvas.create_rectangle(x1, y1, x2, y2, outline="blue", width=2, tags=("selection",))
                self.status_var.set(f"選択: {os.path.basename(self.photo_paths[photo_idx_found])}")
            else:
                self.drag_data = {"item": None, "photo_idx": None, "x": 0, "y": 0}

    def on_motion(self, event):
        """
        マウスがドラッグされているときの処理
        """
        if self.select_mode and self.select_start:
            self.canvas.delete("selection_box")
            x1, y1 = self.select_start
            x2, y2 = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
            self.canvas.create_rectangle(x1, y1, x2, y2, outline="blue", tags="selection_box", dash=(5, 2))
        elif self.drag_data["item"] is not None:
            dx = self.canvas.canvasx(event.x) - self.drag_data["x"]
            dy = self.canvas.canvasy(event.y) - self.drag_data["y"]
            self.canvas.move(self.drag_data["item"], dx, dy)
            self.drag_data["x"] = self.canvas.canvasx(event.x)
            self.drag_data["y"] = self.canvas.canvasy(event.y)
            self.canvas.delete("selection")
            x1, y1, x2, y2 = self.photo_positions[self.drag_data["row"]][self.drag_data["idx"]]
            self.canvas.create_rectangle(x1+dx, y1+dy, x2+dx, y2+dy, outline="blue", width=2, tags=("selection",))

    def on_release(self, event):
        """
        マウスボタンが離されたときの処理
        """
        if self.select_mode and self.select_start:
            self.canvas.delete("selection_box")
            selected_indices = self.get_selected_photo_indices(self.select_start[0], self.select_start[1], self.canvas.canvasx(event.x), self.canvas.canvasy(event.y))
            if selected_indices:
                self.prompt_for_comment(selected_indices)
            self.select_start = None
            self.toggle_select_mode()
        else:
            src_idx = self.drag_data["photo_idx"]
            if src_idx is not None:
                dst_idx = -1
                for r, row in enumerate(self.photo_positions):
                    for i, (x1, y1, x2, y2) in enumerate(row):
                        if x1 <= self.canvas.canvasx(event.x) <= x2 and y1 <= self.canvas.canvasy(event.y) <= y2:
                            dst_idx = sum(self.rows_config[:r]) + i
                            break
                    if dst_idx != -1:
                        break
                
                if dst_idx != -1 and src_idx != dst_idx and 0 <= src_idx < len(self.photos) and 0 <= dst_idx < len(self.photos):
                    self.photos[src_idx], self.photos[dst_idx] = self.photos[dst_idx], self.photos[src_idx]
                    self.photo_paths[src_idx], self.photo_paths[dst_idx] = self.photo_paths[dst_idx], self.photo_paths[src_idx]
                    
                    comment_src = self.photo_comments.get(src_idx)
                    comment_dst = self.photo_comments.get(dst_idx)
                    
                    self.photo_comments.pop(src_idx, None)
                    self.photo_comments.pop(dst_idx, None)
                    if comment_src:
                        self.photo_comments[dst_idx] = comment_src
                    if comment_dst:
                        self.photo_comments[src_idx] = comment_dst
                
                # 自動振り分けの呼び出しを削除
                # ドラッグ＆ドロップ後、写真の順番は変わるが、段の振り分けは手動で行われるべき
                self.update_preview()
                
                if dst_idx != -1 and src_idx != dst_idx:
                    self.status_var.set(f"写真 {src_idx+1} と {dst_idx+1} を入れ替えました")
                else:
                    self.status_var.set("ドラッグ＆ドロップ操作が完了しました")

    def on_right_click(self, event):
        """
        右クリックでメイン写真の選択と、段の移動を行う
        """
        menu = Menu(self.root, tearoff=0)
        
        clicked_photo_info = None
        for r, row in enumerate(self.photo_positions):
            for i, (x1, y1, x2, y2) in enumerate(row):
                if x1 <= self.canvas.canvasx(event.x) <= x2 and y1 <= self.canvas.canvasy(event.y) <= y2:
                    clicked_photo_info = {"row": r, "idx": i, "photo_idx": sum(self.rows_config[:r]) + i}
                    break
            if clicked_photo_info:
                break
        
        if clicked_photo_info:
            current_row = clicked_photo_info["row"]
            
            menu.add_command(label="この写真をメイン画像に設定", command=lambda: self.set_main_photo(clicked_photo_info["row"], clicked_photo_info["idx"]))
            menu.add_separator()
            
            # 「2段目に移動」または「1段目に移動」のオプションを動的に追加
            if current_row == 0 and self.rows_config[0] > 0:
                menu.add_command(label="2段目に移動", command=lambda: self.move_to_row(clicked_photo_info["photo_idx"], 1))
            elif current_row == 1 and self.rows_config[1] > 0:
                menu.add_command(label="1段目に移動", command=lambda: self.move_to_row(clicked_photo_info["photo_idx"], 0))

            menu.tk_popup(event.x_root, event.y_root)

    def set_main_photo(self, row_idx, photo_in_row_idx):
        """
        指定された写真をメイン画像に設定する
        """
        if row_idx < len(self.main_indices) and photo_in_row_idx < self.rows_config[row_idx]:
            self.main_indices[row_idx] = photo_in_row_idx
            self.update_preview()
            self.status_var.set(f"段 {row_idx+1} のメイン写真を変更しました")

    def move_to_row(self, photo_idx, new_row_idx):
        """
        写真を指定された段に移動する
        """
        # 移動元の段と移動先の段を特定する
        current_row_idx = 0 if photo_idx < self.rows_config[0] else 1
        
        if current_row_idx == new_row_idx:
            return

        # 1. 写真、パス、コメントを一時的に取り出す
        photo_to_move = self.photos.pop(photo_idx)
        photo_path_to_move = self.photo_paths.pop(photo_idx)
        comment_to_move = self.photo_comments.get(photo_idx)

        # 2. 移動先の段の先頭に写真を挿入
        if new_row_idx == 0:
            insert_index = 0
        else: # new_row_idx == 1
            insert_index = self.rows_config[0]
            
        self.photos.insert(insert_index, photo_to_move)
        self.photo_paths.insert(insert_index, photo_path_to_move)
        
        # 3. rows_configを更新する
        if current_row_idx == 0:
            self.rows_config[0] -= 1
            self.rows_config[1] += 1
        else:
            self.rows_config[0] += 1
            self.rows_config[1] -= 1
        
        # 4. photo_commentsのキーを再構築する
        # この方法がシンプルでバグが少ない
        new_photo_comments = {}
        for i, photo_path in enumerate(self.photo_paths):
            original_idx = -1
            for k, v in self.photo_comments.items():
                if v == comment_to_move and self.photo_paths[k] == photo_path_to_move:
                    original_idx = k
                    break
            
            # 元のコメントを新しいインデックスで保存
            if original_idx != -1:
                new_photo_comments[i] = self.photo_comments[original_idx]
                
        self.photo_comments = new_photo_comments
        
        # 5. main_indicesを調整する
        # 移動元の段にメイン写真がなければ0にリセット
        if self.rows_config[current_row_idx] == 0:
            self.main_indices[current_row_idx] = 0
        # 移動先の段のメイン写真を調整
        self.main_indices[new_row_idx] = min(self.main_indices[new_row_idx], self.rows_config[new_row_idx] - 1)
        
        self.update_preview()
        self.status_var.set(f"写真を{new_row_idx+1}段目に移動しました")
    
    def on_delete_key(self, event):
        """
        DeleteキーまたはBackspaceキーで選択した写真を削除する
        """
        if self.drag_data["photo_idx"] is not None:
            self.remove_selected_photo()

    def get_selected_photo_indices(self, x1, y1, x2, y2):
        """
        選択範囲内の写真のインデックスを取得する
        """
        selected_indices = []
        for r, row_pos in enumerate(self.photo_positions):
            for i, (px1, py1, px2, py2) in enumerate(row_pos):
                if (max(x1, px1) < min(x2, px2) and max(y1, py1) < min(y2, py2)):
                    photo_idx = sum(self.rows_config[:r]) + i
                    selected_indices.append(photo_idx)
        return selected_indices

    def prompt_for_comment(self, indices):
        """
        選択された写真にコメントを付けるダイアログを表示
        """
        if not indices:
            return
        
        prompt_text = "選択された写真にコメントを入力してください:"
        current_comment = ""
        if indices[0] in self.photo_comments:
            current_comment = self.photo_comments[indices[0]]

        comment = simpledialog.askstring("コメント追加", prompt_text, initialvalue=current_comment)
        if comment is not None:
            for i in indices:
                self.photo_comments[i] = comment
            self.update_preview()
            self.status_var.set(f"{len(indices)}枚の写真にコメントを追加しました")

    # --- 設定保存 / 読み込み ---
    def save_config(self):
        """
        現在のレイアウト設定と写真パスをJSONファイルに保存する
        """
        if not self.photos:
            messagebox.showinfo("警告", "保存する写真がありません")
            return
            
        path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON設定ファイル", "*.json"), ("すべてのファイル", "*.*")]
        )
        if not path:
            return
            
        try:
            config = {
                "rows_config": self.rows_config,
                "main_indices": self.main_indices,
                "main_ratios": [float(s.get()) for s in self.sliders],
                "row_heights": [float(s.get()) for s in self.height_sliders],
                "row_comments": self.row_comments,
                "photo_comments": self.photo_comments,
                "photo_paths": self.photo_paths
            }
            
            with open(path, "w", encoding="utf-8") as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
                
            self.status_var.set(f"設定を保存しました: {path}")
        except Exception as e:
            messagebox.showerror("エラー", f"設定保存中にエラーが発生しました: {str(e)}")

    def load_config(self):
        """
        JSONファイルからレイアウト設定を読み込む
        """
        path = filedialog.askopenfilename(
            filetypes=[("JSON設定ファイル", "*.json"), ("すべてのファイル", "*.*")]
        )
        if not path:
            return
            
        try:
            with open(path, "r", encoding="utf-8") as f:
                config = json.load(f)
                
            self.rows_config = config["rows_config"]
            self.main_indices = config["main_indices"]
            self.row_comments = config.get("row_comments", [""] * len(self.rows_config))
            self.photo_comments = {int(k): v for k, v in config.get("photo_comments", {}).items()}
            
            self.update_sliders()
            
            for i, ratio in enumerate(config.get("main_ratios", [])):
                if i < len(self.sliders):
                    self.sliders[i].set(ratio)
                    
            for i, h in enumerate(config.get("row_heights", [])):
                if i < len(self.height_sliders):
                    self.height_sliders[i].set(h)
            
            for i, comment in enumerate(self.row_comments):
                if i < len(self.comment_entries):
                    self.comment_entries[i].delete(0, tk.END)
                    self.comment_entries[i].insert(0, comment)
            
            self.photos.clear()
            self.photo_paths = []
            loaded = 0
            missing = 0
            
            for f in config.get("photo_paths", []):
                if os.path.exists(f):
                    try:
                        img = Image.open(f)
                        self.photos.append(img)
                        self.photo_paths.append(f)
                        loaded += 1
                    except:
                        missing += 1
                else:
                    missing += 1
            
            self.update_preview()
            
            if missing > 0:
                messagebox.showwarning("警告", f"{missing}枚の写真が見つかりませんでした")
                
            self.status_var.set(f"設定を読み込みました: {loaded}枚の写真")
        except Exception as e:
            messagebox.showerror("エラー", f"設定読み込み中にエラーが発生しました: {str(e)}")

    # --- Excel出力 ---
    def export_excel(self):
        """
        現在のレイアウト設定に基づいてExcelファイルを出力する
        """
        if not self.photos:
            messagebox.showinfo("警告", "写真がありません")
            return
            
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel ファイル", "*.xlsx"), ("すべてのファイル", "*.*")]
        )
        if not path:
            return
            
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "写真レイアウト"
            
            for i in range(1, 100):
                ws.row_dimensions[i].height = 120
            for i in range(1, 30):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = 20
            
            idx = 0
            excel_row = 1
            
            for r_idx, count in enumerate(self.rows_config):
                if idx >= len(self.photos):
                    break
                    
                row_comment_text = self.row_comments[r_idx] if r_idx < len(self.row_comments) else ""
                if row_comment_text:
                    ws.cell(row=excel_row, column=1).value = row_comment_text
                    ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=max(count, 1))
                    cell = ws.cell(row=excel_row, column=1)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    excel_row += 1
                
                row_photos = self.photos[idx:idx+count]
                if not row_photos:
                    break
                    
                main_idx = self.main_indices[r_idx] if r_idx < len(self.main_indices) else 0
                if main_idx >= len(row_photos):
                    main_idx = 0
                
                main_ratio = self.sliders[r_idx].get()
                
                photo_count = len(row_photos)
                if photo_count == 1:
                    main_ratio = 1.0
                
                remaining_ratio = (1 - main_ratio) / (photo_count - 1) if photo_count > 1 else 0
                
                widths = []
                for i in range(photo_count):
                    w = int(1000 * (main_ratio if i == main_idx else remaining_ratio))
                    widths.append(w)
                
                max_height = 0
                for i, img in enumerate(row_photos):
                    h = int(img.height * (widths[i] / img.width))
                    max_height = max(max_height, h)
                
                height_ratio = self.height_sliders[r_idx].get() if r_idx < len(self.height_sliders) else 0.5
                uniform_height = max_height * height_ratio
                
                for i, img in enumerate(row_photos):
                    col = i + 1
                    
                    photo_idx = sum(self.rows_config[:r_idx]) + i
                    if photo_idx in self.photo_comments:
                        comment_cell = ws.cell(row=excel_row, column=col)
                        comment_cell.value = self.photo_comments[photo_idx]
                        comment_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                    width_ratio = main_ratio if i == main_idx else remaining_ratio
                    col_width = int(20 * width_ratio)
                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].width = col_width
                    
                    img_width = int(img.width * (uniform_height / img.height))
                    img_height = int(uniform_height)
                    
                    resized_img = img.resize((img_width, img_height), Image.LANCZOS)
                    
                    with io.BytesIO() as output:
                        resized_img.save(output, format="PNG")
                        output.seek(0)
                        xl_img = XLImage(output)
                        
                        xl_img.width = img_width
                        xl_img.height = img_height
                        
                        ws.add_image(xl_img, f"{col_letter}{excel_row+1}")
                
                ws.row_dimensions[excel_row+1].height = uniform_height * 0.75
                
                excel_row += 2
                idx += count
            
            wb.save(path)
            messagebox.showinfo("完了", f"Excel出力しました: {path}")
            self.status_var.set(f"Excelに出力しました: {path}")
        except Exception as e:
            messagebox.showerror("エラー", f"Excel出力中にエラーが発生しました: {str(e)}")

    # --- ヘルプ ---
    def show_help(self):
        """
        使い方を説明するメッセージボックスを表示する
        """
        msg = (
            "■ 使い方 ■\n\n"
            "【基本操作】\n"
            "1. 画像追加ボタンで写真を選択すると、自動で1段目と2段目に振り分けられます。\n"
            "2. 写真を**ドラッグ＆ドロップ**で、好きな位置に入れ替えられます。段をまたぐことも可能です。\n"
            "3. 右クリックでメイン写真を選択（赤枠表示）。\n"
            "4. Excel出力で保存。\n\n"
            "【詳細設定】\n"
            "・メイン比率: メイン写真の幅比率\n"
            "・高さ倍率: 写真の高さ調整\n"
            "・右クリック: メイン写真の選択と段の移動\n"
            "・設定保存/読み込み: レイアウト設定を保存/読み込み\n"
            "・段コメント: 各段に一言メモを記入\n\n"
            "【写真操作】\n"
            "・左クリック選択 + ドラッグ: 写真を移動して入れ替え\n"
            "・写真の左クリック: 選択状態に（青枠表示）\n"
            "・「写真入れ替え」ボタン: 写真の番号を指定して入れ替え\n"
            "・「コメント追加」ボタン: マウスで複数写真を選択してコメントをまとめて追加\n"
            "・写真選択 + Delete/Backspaceキー: 選択した写真を削除"
        )
        messagebox.showinfo("ヘルプ", msg)


if __name__ == "__main__":
    root = tk.Tk()
    app = PhotoLayoutApp(root)
    root.mainloop()

