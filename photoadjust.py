# -*- coding: utf-8 -*-

import os
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from PIL import Image, ImageTk
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.cell.cell import Cell
import io
import json

class PhotoLayoutApp:
    """
    Excelに写真をレイアウトするGUIアプリケーション
    """
    def __init__(self, root):
        self.root = root
        self.root.title("Excel写真レイアウトアプリ 最終版")
        
        # 状態変数
        self.photos = []
        self.photo_paths = []
        self.rows_config = [3, 4]  # 各段の列数
        self.main_indices = [0] * len(self.rows_config)
        self.row_comments = [""] * len(self.rows_config) # 各段のコメントを保持
        self.photo_comments = {} # 個別写真のコメントを保持 (例: {0: "コメント1", 2: "コメント2"})

        # UI表示用の変数とオブジェクト
        self.photo_positions = []
        self.sliders = []
        self.height_sliders = []
        self.row_entries = []
        self.comment_entries = []
        self.tk_images = []  # TkinterのPhotoImageオブジェクトへの参照を保持
        
        # ドラッグ＆ドロップ用の状態
        self.drag_data = {"item": None, "row": None, "idx": None, "photo_idx": None, "x": 0, "y": 0}
        
        # 範囲選択モードの状態
        self.select_mode = False
        self.select_start = None
        
        self.setup_ui()
        self.update_sliders()
        
        # ウィンドウサイズ設定
        self.root.geometry("900x700")
        self.root.update()
        self.update_preview()

    # --- UI構築 ---
    def setup_ui(self):
        """
        GUIの各ウィジェットを構築する
        """
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill="both", expand=True)
        
        # ボタンフレーム
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=5, padx=5)
        
        tk.Button(btn_frame, text="画像追加", command=self.add_photos, width=10).pack(side="left", padx=2)
        tk.Button(btn_frame, text="画像削除", command=self.remove_selected_photo, width=10).pack(side="left", padx=2)
        tk.Button(btn_frame, text="Excel出力", command=self.export_excel, width=10).pack(side="left", padx=2)
        tk.Button(btn_frame, text="設定保存", command=self.save_config, width=10).pack(side="left", padx=2)
        tk.Button(btn_frame, text="設定読み込み", command=self.load_config, width=10).pack(side="left", padx=2)
        
        # コントロールフレーム
        control_frame = tk.Frame(main_frame)
        control_frame.pack(fill="x", pady=5, padx=5)
        
        tk.Label(control_frame, text="段数:").pack(side="left")
        self.row_count_var = tk.StringVar(value=str(len(self.rows_config)))
        row_count_entry = tk.Entry(control_frame, textvariable=self.row_count_var, width=3)
        row_count_entry.pack(side="left", padx=5)
        
        tk.Button(control_frame, text="段数更新", command=self.update_row_count).pack(side="left", padx=2)
        tk.Button(control_frame, text="列数更新", command=self.update_rows_config).pack(side="left", padx=2)
        tk.Button(control_frame, text="写真入れ替え", command=self.swap_photos_dialog, width=12).pack(side="left", padx=2)
        tk.Button(control_frame, text="コメント追加", command=self.toggle_select_mode, width=12).pack(side="left", padx=2)
        tk.Button(control_frame, text="ヘルプ", command=self.show_help).pack(side="left", padx=2)
        
        # キャンバスとスクロールバー
        canvas_frame = tk.Frame(main_frame)
        canvas_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.canvas = tk.Canvas(canvas_frame, bg="#eee")
        self.canvas.pack(fill="both", expand=True)
        
        self.canvas.bind("<ButtonPress-1>", self.on_press)
        self.canvas.bind("<B1-Motion>", self.on_motion)
        self.canvas.bind("<ButtonRelease-1>", self.on_release)
        self.canvas.bind("<Button-3>", self.on_right_click)
        self.canvas.bind("<Configure>", lambda e: self.update_preview())

        self.slider_frame = tk.Frame(main_frame)
        self.slider_frame.pack(fill="x", padx=5, pady=5)
        
        self.status_var = tk.StringVar()
        status_bar = tk.Label(main_frame, textvariable=self.status_var, anchor="w", bd=1, relief=tk.SUNKEN)
        status_bar.pack(fill="x", side="bottom")

    # --- スライダー更新 ---
    def update_sliders(self):
        """
        段数設定に合わせてスライダーUIを更新する
        """
        # 既存のスライダーを全て削除
        for w in self.slider_frame.winfo_children():
            w.destroy()
        self.sliders.clear()
        self.height_sliders.clear()
        self.row_entries.clear()
        self.comment_entries.clear()
        
        # 新しいスライダーを作成
        max_cols = 4 # 1行に表示するパネルの最大数
        for i in range(len(self.rows_config)):
            row = i // max_cols
            col = i % max_cols
            
            frame = tk.LabelFrame(self.slider_frame, text=f"段 {i+1} 設定")
            frame.grid(row=row, column=col, padx=5, pady=5, sticky="nw")
            
            # メイン比率スライダー
            tk.Label(frame, text="メイン比率").grid(row=0, column=0, sticky="w")
            slider = tk.Scale(frame, from_=0.1, to=0.9, resolution=0.01, orient="horizontal",
                             command=lambda e, idx=i: self.update_preview(), length=100)
            slider.set(0.5)
            slider.grid(row=0, column=1)
            self.sliders.append(slider)

            # 高さ倍率スライダー
            tk.Label(frame, text="高さ倍率").grid(row=1, column=0, sticky="w")
            h_slider = tk.Scale(frame, from_=0.5, to=2.0, resolution=0.01, orient="horizontal",
                                command=lambda e, idx=i: self.update_preview(), length=100)
            h_slider.set(1.0)
            h_slider.grid(row=1, column=1)
            self.height_sliders.append(h_slider)

            # 列数入力欄
            tk.Label(frame, text="列数").grid(row=2, column=0, sticky="w")
            e = tk.Entry(frame, width=3)
            e.insert(0, str(self.rows_config[i]))
            e.grid(row=2, column=1, sticky="w")
            self.row_entries.append(e)

            # 段ごとのコメント入力欄
            tk.Label(frame, text="段コメント").grid(row=3, column=0, sticky="w")
            comment_var = tk.StringVar(value=self.row_comments[i] if i < len(self.row_comments) else "")
            comment_entry = tk.Entry(frame, width=20, textvariable=comment_var)
            comment_entry.grid(row=3, column=1, sticky="w")
            self.comment_entries.append(comment_entry)
            
            comment_entry.bind("<KeyRelease>", lambda e, idx=i: self.update_row_comment(idx, comment_entry.get()))
            
    def update_row_comment(self, row_idx, text):
        """
        段コメント入力欄の値が変更されたときに呼び出される
        """
        while len(self.row_comments) <= row_idx:
            self.row_comments.append("")
        self.row_comments[row_idx] = text
        self.update_preview()

    def toggle_select_mode(self):
        """
        コメント追加モードのON/OFFを切り替える
        """
        self.select_mode = not self.select_mode
        if self.select_mode:
            self.status_var.set("コメント追加モード: 写真を範囲選択してください")
            self.canvas.config(cursor="cross")
            self.drag_data = {"item": None, "row": None, "idx": None, "photo_idx": None, "x": 0, "y": 0}
        else:
            self.status_var.set("通常モード")
            self.canvas.config(cursor="arrow")
            self.canvas.delete("selection_box")

    def update_row_count(self):
        """
        段数の入力値に基づいて段数を更新する
        """
        try:
            count = int(self.row_count_var.get())
            if count < 1:
                raise ValueError("段数は1以上である必要があります")
            
            # rows_config, main_indices, row_commentsを更新
            if count > len(self.rows_config):
                self.rows_config.extend([3] * (count - len(self.rows_config)))
                self.main_indices.extend([0] * (count - len(self.main_indices)))
                self.row_comments.extend([""] * (count - len(self.row_comments)))
            else:
                self.rows_config = self.rows_config[:count]
                self.main_indices = self.main_indices[:count]
                self.row_comments = self.row_comments[:count]
                
            self.update_sliders()
            self.update_preview()
        except ValueError as e:
            messagebox.showerror("エラー", str(e))

    def update_rows_config(self):
        """
        各段の列数の入力値に基づいて設定を更新する
        """
        for i, e in enumerate(self.row_entries):
            try:
                val = int(e.get())
                if val < 1:
                    messagebox.showwarning("警告", f"段 {i+1} の列数は1以上にしてください")
                    val = 1
                self.rows_config[i] = val
            except ValueError:
                messagebox.showwarning("警告", f"段 {i+1} の列数が無効です")
                
        # メイン写真のインデックスを調整
        for i, main_idx in enumerate(self.main_indices):
            if main_idx >= self.rows_config[i]:
                self.main_indices[i] = 0
                
        self.update_preview()

    # --- 写真追加・削除 ---
    def add_photos(self):
        """
        ファイルダイアログから画像ファイルを選択し、リストに追加する
        """
        files = filedialog.askopenfilenames(filetypes=[
            ("画像ファイル", "*.png;*.jpg;*.jpeg;*.bmp;*.gif;*.tiff"),
            ("すべてのファイル", "*.*")
        ])
        
        added = 0
        for f in files:
            try:
                img = Image.open(f)
                self.photos.append(img)
                self.photo_paths.append(f)
                added += 1
            except Exception as e:
                self.status_var.set(f"画像読み込み失敗: {os.path.basename(f)} ({str(e)})")
        
        if added > 0:
            self.status_var.set(f"{added}枚の画像を追加しました")
            self.update_preview()

    def remove_selected_photo(self):
        """
        選択中の写真をリストから削除する
        """
        if not self.drag_data["photo_idx"] is None:
            idx = self.drag_data["photo_idx"]
            if 0 <= idx < len(self.photos):
                del self.photos[idx]
                del self.photo_paths[idx]
                # ドラッグデータをリセット
                self.drag_data = {"item": None, "row": None, "idx": None, "photo_idx": None, "x": 0, "y": 0}
                self.update_preview()
                self.status_var.set("選択した写真を削除しました")
        else:
            messagebox.showinfo("情報", "削除する写真を選択してください")

    def swap_photos_dialog(self):
        """
        番号指定で写真の位置を入れ替えるダイアログを表示する
        """
        if len(self.photos) < 2:
            messagebox.showinfo("情報", "写真を2枚以上追加してください。")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("写真入れ替え")
        dialog.geometry("300x150")

        tk.Label(dialog, text="入れ替えたい写真の番号を入力してください。").pack(pady=5)
        tk.Label(dialog, text="(例: 1, 3)").pack()

        frame = tk.Frame(dialog)
        frame.pack(pady=5)
        
        tk.Label(frame, text="番号1:").pack(side="left", padx=5)
        entry1 = tk.Entry(frame, width=5)
        entry1.pack(side="left", padx=5)
        
        tk.Label(frame, text="番号2:").pack(side="left", padx=5)
        entry2 = tk.Entry(frame, width=5)
        entry2.pack(side="left", padx=5)

        def perform_swap():
            try:
                idx1 = int(entry1.get()) - 1
                idx2 = int(entry2.get()) - 1

                if not (0 <= idx1 < len(self.photos) and 0 <= idx2 < len(self.photos)):
                    messagebox.showerror("エラー", "無効な写真番号です。")
                    return

                # 写真を入れ替え
                self.photos[idx1], self.photos[idx2] = self.photos[idx2], self.photos[idx1]
                self.photo_paths[idx1], self.photo_paths[idx2] = self.photo_paths[idx2], self.photo_paths[idx1]
                
                self.update_preview()
                self.status_var.set(f"写真 {idx1+1} と {idx2+1} を入れ替えました")
                dialog.destroy()
            except ValueError:
                messagebox.showerror("エラー", "有効な数字を入力してください。")

        tk.Button(dialog, text="入れ替え実行", command=perform_swap).pack(pady=10)


    # --- Canvas描画 ---
    def update_preview(self):
        """
        設定に基づいてCanvas上に写真のプレビューを描画する
        今回は画面に収まるように高さを自動調整
        """
        self.canvas.delete("all")
        self.photo_positions.clear()
        self.tk_images = []
        
        if not self.photos:
            self.status_var.set("写真を追加してください")
            self.canvas.configure(scrollregion=(0, 0, 800, 100))
            return
            
        canvas_w = self.canvas.winfo_width() - 20
        canvas_h = self.canvas.winfo_height() - 20
        if canvas_w < 100 or canvas_h < 100:
            canvas_w = 800
            canvas_h = 600
        
        y = 10
        total_photos = len(self.photos)
        
        # 画面に収まるように全体のスケールを計算
        temp_y = 10
        total_row_heights = []
        for r_idx, count in enumerate(self.rows_config):
            if sum(self.rows_config[:r_idx]) >= total_photos:
                break
            
            row_photos_count = min(count, total_photos - sum(self.rows_config[:r_idx]))
            if row_photos_count == 0:
                continue

            main_ratio = self.sliders[r_idx].get()
            remaining_ratio = (1 - main_ratio) / (row_photos_count - 1) if row_photos_count > 1 else 0

            max_h_temp = 0
            for i in range(row_photos_count):
                w = int(canvas_w * (main_ratio if i == self.main_indices[r_idx] else remaining_ratio))
                h = int(self.photos[i].height * (w / self.photos[i].width) * self.height_sliders[r_idx].get())
                max_h_temp = max(max_h_temp, h)
            total_row_heights.append(max_h_temp)
            temp_y += max_h_temp + 30
        
        total_height = sum(total_row_heights) + (len(total_row_heights) * 30)
        scale = canvas_h / total_height if total_height > canvas_h else 1.0
        
        max_width = 0
        idx = 0
        for r_idx, count in enumerate(self.rows_config):
            if idx >= len(self.photos):
                break
            
            row_photos = self.photos[idx:idx + count]
            if not row_photos:
                break
            
            # 段コメントの描画
            row_comment_text = self.row_comments[r_idx] if r_idx < len(self.row_comments) else ""
            if row_comment_text:
                self.canvas.create_text(canvas_w // 2, y, text=row_comment_text, anchor="n", fill="black", font=("Arial", 12, "bold"))
                y += 25 # コメント分のスペースを空ける

            main_idx = self.main_indices[r_idx] if self.main_indices[r_idx] < len(row_photos) else 0
            main_ratio = self.sliders[r_idx].get()
            
            photo_count = len(row_photos)
            if photo_count == 1:
                main_ratio = 1.0
                
            remaining_ratio = (1 - main_ratio) / (photo_count - 1) if photo_count > 1 else 0
            
            heights = []
            for i, img in enumerate(row_photos):
                w = int(canvas_w * (main_ratio if i == main_idx else remaining_ratio))
                h = int(img.height * (w / img.width) * self.height_sliders[r_idx].get())
                heights.append(h)
                
            max_h = max(heights) if heights else 0
            
            x = 0
            row_positions = []
            for i, img in enumerate(row_photos):
                w = int(canvas_w * (main_ratio if i == main_idx else remaining_ratio) * scale)
                h = int(img.height * (w / img.width) * self.height_sliders[r_idx].get())
                
                max_width = max(max_width, x + w)
                
                try:
                    resized_img = img.resize((w, h), Image.LANCZOS)
                    tk_img = ImageTk.PhotoImage(resized_img)
                    self.tk_images.append(tk_img)
                    
                    self.canvas.create_image(x, y, anchor="nw", image=tk_img,
                                             tags=(f"photo_{r_idx}_{i}",))
                    
                    # メイン写真に赤枠を表示
                    if i == main_idx:
                        self.canvas.create_rectangle(x, y, x + w, y + h, outline="red", width=3,
                                                     tags=(f"main_{r_idx}",))
                    
                    # 個別写真コメントの描画
                    photo_idx = sum(min(self.rows_config[j], len(self.photos) - sum(self.rows_config[:j]))
                                     for j in range(r_idx)) + i
                    if photo_idx in self.photo_comments:
                        self.canvas.create_text(x + w//2, y + 10, text=self.photo_comments[photo_idx],
                                                anchor="n", fill="blue", font=("Arial", 10))

                    # ファイル名テキストと番号を表示
                    if photo_idx < len(self.photo_paths):
                        filename = os.path.basename(self.photo_paths[photo_idx])
                        text_to_show = f"({photo_idx+1}) {filename[:20]}"
                        self.canvas.create_text(x + w//2, y + h + 5, text=text_to_show,
                                                 anchor="n", fill="black", tags=(f"label_{r_idx}_{i}",))
                    
                    row_positions.append((x, y, x + w, y + h))
                    x += w
                except Exception as e:
                    self.status_var.set(f"画像表示エラー: {str(e)}")
            
            self.photo_positions.append(row_positions)
            y += max_h * scale + 30
            idx += count
        
        # スクロール領域を更新
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.status_var.set(f"表示中: {len(self.photos)}枚の写真")

    # --- ドラッグ＆ドロップ操作 ---
    def on_press(self, event):
        """
        マウスボタンが押されたときの処理
        """
        self.canvas.delete("selection")
        if self.select_mode:
            self.select_start = (event.x, event.y)
            self.canvas.delete("selection_box")
        else:
            self.drag_data = {"item": None, "row": None, "idx": None, "photo_idx": None, "x": event.x, "y": event.y}
            for r, row in enumerate(self.photo_positions):
                for i, (x1, y1, x2, y2) in enumerate(row):
                    if x1 <= event.x <= x2 and y1 <= event.y <= y2:
                        photo_idx = sum(len(self.photo_positions[j]) for j in range(r)) + i
                        if photo_idx < len(self.photos):
                            self.drag_data.update({
                                "item": self.canvas.find_withtag(f"photo_{r}_{i}"),
                                "row": r,
                                "idx": i,
                                "photo_idx": photo_idx,
                                "x": event.x,
                                "y": event.y
                            })
                            self.canvas.create_rectangle(x1, y1, x2, y2, outline="blue", width=2, tags=("selection",))
                            self.status_var.set(f"選択: {os.path.basename(self.photo_paths[photo_idx])}")
                            return

    def on_motion(self, event):
        """
        マウスがドラッグされているときの処理
        """
        if self.select_mode and self.select_start:
            self.canvas.delete("selection_box")
            x1, y1 = self.select_start
            x2, y2 = event.x, event.y
            self.canvas.create_rectangle(x1, y1, x2, y2, outline="blue", tags="selection_box", dash=(5, 2))
        elif not self.drag_data["item"]:
            return
        else:
            dx = event.x - self.drag_data["x"]
            dy = event.y - self.drag_data["y"]
            self.drag_data["x"] = event.x
            self.drag_data["y"] = event.y
            self.canvas.delete("selection")
            r = self.drag_data["row"]
            i = self.drag_data["idx"]
            if r < len(self.photo_positions) and i < len(self.photo_positions[r]):
                x1, y1, x2, y2 = self.photo_positions[r][i]
                self.canvas.create_rectangle(x1+dx, y1+dy, x2+dx, y2+dy, outline="blue", width=2, tags=("selection",))

    def on_release(self, event):
        """
        マウスボタンが離されたときの処理
        """
        if self.select_mode and self.select_start:
            self.canvas.delete("selection_box")
            selected_indices = self.get_selected_photo_indices(self.select_start[0], self.select_start[1], event.x, event.y)
            if selected_indices:
                self.prompt_for_comment(selected_indices)
            self.select_start = None
            self.toggle_select_mode()
        elif not self.drag_data["item"]:
            return
        else:
            src_idx = self.drag_data["photo_idx"]
            for r, row in enumerate(self.photo_positions):
                for i, (x1, y1, x2, y2) in enumerate(row):
                    if x1 <= event.x <= x2 and y1 <= event.y <= y2:
                        dst_idx = sum(len(self.photo_positions[j]) for j in range(r)) + i
                        if src_idx != dst_idx and 0 <= src_idx < len(self.photos) and 0 <= dst_idx < len(self.photos):
                            self.photos[src_idx], self.photos[dst_idx] = self.photos[dst_idx], self.photos[src_idx]
                            self.photo_paths[src_idx], self.photo_paths[dst_idx] = self.photo_paths[dst_idx], self.photo_paths[src_idx]
                            
                            # コメントも一緒に移動
                            comment_src = self.photo_comments.get(src_idx, None)
                            comment_dst = self.photo_comments.get(dst_idx, None)
                            if comment_src:
                                self.photo_comments[dst_idx] = comment_src
                            else:
                                self.photo_comments.pop(dst_idx, None)
                            if comment_dst:
                                self.photo_comments[src_idx] = comment_dst
                            else:
                                self.photo_comments.pop(src_idx, None)

                            self.update_preview()
                            self.status_var.set(f"写真 {src_idx+1} と {dst_idx+1} を入れ替えました")
                            break
            self.canvas.delete("selection")
            self.drag_data = {"item": None, "row": None, "idx": None, "photo_idx": None, "x": 0, "y": 0}

    def on_right_click(self, event):
        """
        右クリックでメイン写真を選択する
        """
        for r, row in enumerate(self.photo_positions):
            for i, (x1, y1, x2, y2) in enumerate(row):
                if x1 <= event.x <= x2 and y1 <= event.y <= y2:
                    self.main_indices[r] = i
                    self.update_preview()
                    self.status_var.set(f"段 {r+1} のメイン写真を変更しました")
                    return
    
    def get_selected_photo_indices(self, x1, y1, x2, y2):
        """
        選択範囲内の写真のインデックスを取得する
        """
        selected_indices = []
        for r, row_pos in enumerate(self.photo_positions):
            for i, (px1, py1, px2, py2) in enumerate(row_pos):
                if (max(x1, px1) < min(x2, px2) and max(y1, py1) < min(y2, py2)):
                    photo_idx = sum(len(self.photo_positions[j]) for j in range(r)) + i
                    selected_indices.append(photo_idx)
        return selected_indices

    def prompt_for_comment(self, indices):
        """
        選択された写真にコメントを付けるダイアログを表示
        """
        if not indices:
            return
        
        prompt_text = "選択された写真にコメントを入力してください:"
        current_comment = ""
        # 最初の写真のコメントがあれば表示
        if indices[0] in self.photo_comments:
            current_comment = self.photo_comments[indices[0]]

        comment = simpledialog.askstring("コメント追加", prompt_text, initialvalue=current_comment)
        if comment is not None:
            # 既存のコメントをクリアしてから新しいコメントを追加
            for i in indices:
                self.photo_comments[i] = comment
            self.update_preview()
            self.status_var.set(f"{len(indices)}枚の写真にコメントを追加しました")

    # --- 設定保存 / 読み込み ---
    def save_config(self):
        """
        現在のレイアウト設定と写真パスをJSONファイルに保存する
        """
        if not self.photos:
            messagebox.showinfo("警告", "保存する写真がありません")
            return
            
        path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON設定ファイル", "*.json"), ("すべてのファイル", "*.*")]
        )
        if not path:
            return
            
        try:
            config = {
                "rows_config": self.rows_config,
                "main_indices": self.main_indices,
                "main_ratios": [float(s.get()) for s in self.sliders],
                "row_heights": [float(s.get()) for s in self.height_sliders],
                "row_comments": self.row_comments,
                "photo_comments": self.photo_comments,
                "photo_paths": self.photo_paths
            }
            
            with open(path, "w", encoding="utf-8") as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
                
            self.status_var.set(f"設定を保存しました: {path}")
        except Exception as e:
            messagebox.showerror("エラー", f"設定保存中にエラーが発生しました: {str(e)}")

    def load_config(self):
        """
        JSONファイルからレイアウト設定を読み込む
        """
        path = filedialog.askopenfilename(
            filetypes=[("JSON設定ファイル", "*.json"), ("すべてのファイル", "*.*")]
        )
        if not path:
            return
            
        try:
            with open(path, "r", encoding="utf-8") as f:
                config = json.load(f)
                
            self.rows_config = config["rows_config"]
            self.main_indices = config["main_indices"]
            self.row_comments = config.get("row_comments", [""] * len(self.rows_config))
            self.photo_comments = config.get("photo_comments", {})
            
            self.row_count_var.set(str(len(self.rows_config)))
            self.update_sliders()
            
            # スライダーとコメントの値を設定
            for i, ratio in enumerate(config.get("main_ratios", [])):
                if i < len(self.sliders):
                    self.sliders[i].set(ratio)
                    
            for i, h in enumerate(config.get("row_heights", [])):
                if i < len(self.height_sliders):
                    self.height_sliders[i].set(h)
            
            for i, comment in enumerate(self.row_comments):
                if i < len(self.comment_entries):
                    self.comment_entries[i].delete(0, tk.END)
                    self.comment_entries[i].insert(0, comment)
            
            # 写真を読み込む
            self.photos.clear()
            self.photo_paths = []
            loaded = 0
            missing = 0
            
            for f in config.get("photo_paths", []):
                if os.path.exists(f):
                    try:
                        img = Image.open(f)
                        self.photos.append(img)
                        self.photo_paths.append(f)
                        loaded += 1
                    except:
                        missing += 1
                else:
                    missing += 1
            
            self.update_preview()
            
            if missing > 0:
                messagebox.showwarning("警告", f"{missing}枚の写真が見つかりませんでした")
                
            self.status_var.set(f"設定を読み込みました: {loaded}枚の写真")
        except Exception as e:
            messagebox.showerror("エラー", f"設定読み込み中にエラーが発生しました: {str(e)}")

    # --- Excel出力 ---
    def export_excel(self):
        """
        現在のレイアウト設定に基づいてExcelファイルを出力する
        """
        if not self.photos:
            messagebox.showinfo("警告", "写真がありません")
            return
            
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel ファイル", "*.xlsx"), ("すべてのファイル", "*.*")]
        )
        if not path:
            return
            
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "写真レイアウト"
            
            # デフォルトの行の高さと列の幅を設定
            for i in range(1, 100):
                ws.row_dimensions[i].height = 120
            for i in range(1, 30):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = 20
            
            idx = 0
            excel_row = 1
            
            for r_idx, count in enumerate(self.rows_config):
                if idx >= len(self.photos):
                    break
                    
                # 段コメントのセル
                row_comment_text = self.row_comments[r_idx] if r_idx < len(self.row_comments) else ""
                if row_comment_text:
                    ws.cell(row=excel_row, column=1).value = row_comment_text
                    
                    # セルを結合
                    ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=count)
                    
                    # 文字位置を調整
                    cell = ws.cell(row=excel_row, column=1)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    excel_row += 1 # コメント行分、行をずらす
                
                # 個別コメントと写真の出力
                row_photos = self.photos[idx:idx+count]
                if not row_photos:
                    break
                    
                main_idx = self.main_indices[r_idx] if self.main_indices[r_idx] < len(row_photos) else 0
                main_ratio = self.sliders[r_idx].get()
                
                photo_count = len(row_photos)
                if photo_count == 1:
                    main_ratio = 1.0
                    
                remaining_ratio = (1 - main_ratio) / (photo_count - 1) if photo_count > 1 else 0
                
                # 写真をワークシートに追加
                for i, img in enumerate(row_photos):
                    col = i + 1
                    
                    # 個別コメントのセル
                    photo_idx = sum(min(self.rows_config[j], len(self.photos) - sum(self.rows_config[:j]))
                                     for j in range(r_idx)) + i
                    if photo_idx in self.photo_comments:
                        comment_cell = ws.cell(row=excel_row, column=col)
                        comment_cell.value = self.photo_comments[photo_idx]
                        comment_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                    # 列幅を計算
                    width_ratio = main_ratio if i == main_idx else remaining_ratio
                    col_width = int(100 * width_ratio)
                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].width = col_width
                    
                    # 画像のサイズを計算
                    img_width = int(300 * width_ratio)
                    img_height = int(img.height * (img_width / img.width) * self.height_sliders[r_idx].get())
                    
                    resized_img = img.resize((img_width, img_height), Image.LANCZOS)
                    
                    # 画像をExcelに追加
                    with io.BytesIO() as output:
                        resized_img.save(output, format="PNG")
                        output.seek(0)
                        xl_img = XLImage(output)
                        
                        xl_img.width = img_width
                        xl_img.height = img_height
                        
                        ws.add_image(xl_img, f"{col_letter}{excel_row+1}") # コメント行を考慮して1行ずらす
                
                # 写真の高さに合わせた行の高さを設定
                max_h = max([img.height for img in row_photos]) if row_photos else 0
                ws.row_dimensions[excel_row+1].height = max_h * 0.75 # openpyxlの単位に変換
                
                excel_row += 1 + (max(img_height for img in row_photos) // 100)
                idx += count
            
            wb.save(path)
            messagebox.showinfo("完了", f"Excel出力しました: {path}")
            self.status_var.set(f"Excelに出力しました: {path}")
        except Exception as e:
            messagebox.showerror("エラー", f"Excel出力中にエラーが発生しました: {str(e)}")

    # --- ヘルプ ---
    def show_help(self):
        """
        使い方を説明するメッセージボックスを表示する
        """
        msg = (
            "■ 使い方 ■\n\n"
            "【基本操作】\n"
            "1. 画像追加ボタンで写真を追加\n"
            "2. 各段の列数・メイン比率・高さ倍率を設定\n"
            "3. 右クリックでメイン写真選択（赤枠表示）\n"
            "4. 写真のドラッグで位置入れ替え可能\n"
            "5. Excel出力で保存\n\n"
            "【詳細設定】\n"
            "・「段数更新」ボタン: 段数を変更\n"
            "・「列数更新」ボタン: 各段の列数を変更\n"
            "・メイン比率: メイン写真の幅比率\n"
            "・高さ倍率: 写真の高さ調整\n"
            "・右クリック: メイン写真を選択\n"
            "・設定保存/読み込み: レイアウト設定を保存/読み込み\n"
            "・段コメント: 各段に一言メモを記入\n\n"
            "【写真操作】\n"
            "・左クリック選択 + ドラッグ: 写真を移動して入れ替え\n"
            "・「写真入れ替え」ボタン: 写真の番号を指定して入れ替え\n"
            "・「コメント追加」ボタン: マウスで複数写真を選択してコメントをまとめて追加\n"
            "・写真選択 + 削除ボタン: 選択した写真を削除"
        )
        messagebox.showinfo("ヘルプ", msg)


if __name__ == "__main__":
    root = tk.Tk()
    app = PhotoLayoutApp(root)
    root.mainloop()
